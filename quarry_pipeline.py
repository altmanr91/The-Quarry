import json
import os
import sys
from datetime import date
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from comps_writer import append_articles, _purge_no_basis
from contacts_writer import upsert_contacts

load_dotenv()

HANDOFF_URL   = 'https://altmanr91.github.io/CRE-News-Reader/articles_handoff.json'
COMPS_FILE    = Path('comps.xlsx')
CONTACTS_FILE = Path('contacts.xlsx')
ONEDRIVE_DIR  = 'Documents/AI Tools/The Quarry'
TOKEN_URL     = 'https://login.microsoftonline.com/consumers/oauth2/v2.0/token'
GRAPH_URL     = 'https://graph.microsoft.com/v1.0'


def _fetch_handoff() -> dict:
    resp = requests.get(HANDOFF_URL, timeout=30)
    resp.raise_for_status()
    return resp.json()


def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _get_access_token(client_id: str, client_secret: str, refresh_token: str) -> str:
    resp = requests.post(TOKEN_URL, data={
        'grant_type':    'refresh_token',
        'refresh_token': refresh_token,
        'client_id':     client_id,
        'client_secret': client_secret,
        'scope':         'offline_access Files.ReadWrite',
    })
    resp.raise_for_status()
    return resp.json()['access_token']


def _od_headers(token: str) -> dict:
    return {'Authorization': f'Bearer {token}'}


def _od_download(token: str, filename: str, dest: Path) -> bool:
    url  = f'{GRAPH_URL}/me/drive/root:/{ONEDRIVE_DIR}/{filename}:/content'
    resp = requests.get(url, headers=_od_headers(token))
    if resp.status_code == 404:
        return False
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return True


def _od_upload(token: str, filename: str, path: Path) -> requests.Response:
    url  = f'{GRAPH_URL}/me/drive/root:/{ONEDRIVE_DIR}/{filename}:/content'
    data = path.read_bytes()
    return requests.put(url, headers={**_od_headers(token), 'Content-Type': 'application/octet-stream'}, data=data)


def _od_list_folder(token: str) -> list:
    url  = f'{GRAPH_URL}/me/drive/root:/{ONEDRIVE_DIR}:/children'
    resp = requests.get(url, headers=_od_headers(token))
    if resp.status_code == 404:
        return []
    resp.raise_for_status()
    return resp.json().get('value', [])


def _od_delete(token: str, item_id: str) -> None:
    requests.delete(f'{GRAPH_URL}/me/drive/items/{item_id}', headers=_od_headers(token))


def _download_from_onedrive(comps_path: Path, contacts_path: Path) -> None:
    client_id     = os.getenv('ONEDRIVE_CLIENT_ID')
    client_secret = os.getenv('ONEDRIVE_CLIENT_SECRET')
    refresh_token = os.getenv('ONEDRIVE_REFRESH_TOKEN')

    if not all([client_id, client_secret, refresh_token]):
        print('  [onedrive] Credentials not set — skipping download')
        return

    try:
        token = _get_access_token(client_id, client_secret, refresh_token)
    except Exception as e:
        print(f'  [onedrive] Could not get access token — skipping download: {e}')
        return

    # Absorb any pending comps file saved when comps.xlsx was locked on a prior run
    comps_downloaded = False
    items = _od_list_folder(token)
    pending = sorted(
        [f for f in items if f['name'].startswith('comps_pending_') and f['name'].endswith('.xlsx')],
        key=lambda f: f['name'],
    )
    if pending:
        latest = pending[-1]
        print(f'  [onedrive] Absorbing {latest["name"]} as comps base')
        resp = requests.get(f'{GRAPH_URL}/me/drive/items/{latest["id"]}/content', headers=_od_headers(token))
        resp.raise_for_status()
        comps_path.write_bytes(resp.content)
        comps_downloaded = True
        for p in pending:
            _od_delete(token, p['id'])
            print(f'  [onedrive] Deleted {p["name"]}')

    for path, skip in ((comps_path, comps_downloaded), (contacts_path, False)):
        if skip:
            continue
        found = _od_download(token, path.name, path)
        if found:
            print(f'  [onedrive] Downloaded {path.name}')
        else:
            print(f'  [onedrive] {path.name} not found — will create fresh')


def _upload_to_onedrive(comps_path: Path, contacts_path: Path, date_str: str) -> None:
    client_id     = os.getenv('ONEDRIVE_CLIENT_ID')
    client_secret = os.getenv('ONEDRIVE_CLIENT_SECRET')
    refresh_token = os.getenv('ONEDRIVE_REFRESH_TOKEN')

    if not all([client_id, client_secret, refresh_token]):
        print('  [onedrive] Credentials not set — skipping upload')
        return

    token = _get_access_token(client_id, client_secret, refresh_token)

    for path in (comps_path, contacts_path):
        resp = _od_upload(token, path.name, path)
        if resp.status_code == 423 and path == comps_path:
            pending_name = f'comps_pending_{date_str}.xlsx'
            r2 = _od_upload(token, pending_name, path)
            r2.raise_for_status()
            print(f'  [onedrive] comps.xlsx locked — saved as {pending_name}, will merge on next run')
            continue
        resp.raise_for_status()
        print(f'  [onedrive] Uploaded {path.name}')


def main() -> None:
    print('Fetching handoff JSON...')
    try:
        handoff = _fetch_handoff()
    except Exception as e:
        print(f'ERROR: Could not fetch handoff: {e}')
        sys.exit(1)

    date_str = handoff.get('date') or date.today().isoformat()
    articles = handoff.get('articles') or []
    print(f'  {len(articles)} articles for {date_str}')

    print('Downloading workbooks from OneDrive...')
    _download_from_onedrive(COMPS_FILE, CONTACTS_FILE)

    print('Loading workbooks...')
    comps_wb    = _load_or_create(COMPS_FILE)
    contacts_wb = _load_or_create(CONTACTS_FILE)

    _purge_no_basis(comps_wb)

    print('Writing comps...')
    counts = append_articles(date_str, articles, comps_wb)
    print(f'  Sales: {counts["sales"]}, Leases: {counts["leases"]}, Loans: {counts["loans"]}, SFR: {counts["sfr"]}')

    print('Upserting contacts...')
    new_contacts = upsert_contacts(date_str, articles, contacts_wb)
    print(f'  New contacts: {new_contacts}')

    print('Saving workbooks...')
    comps_wb.save(COMPS_FILE)
    contacts_wb.save(CONTACTS_FILE)

    print('Uploading to OneDrive...')
    _upload_to_onedrive(COMPS_FILE, CONTACTS_FILE, date_str)

    print('Done.')


if __name__ == '__main__':
    main()
