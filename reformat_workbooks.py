"""
One-time script: apply number formats to all existing data rows in comps.xlsx.
Downloads from OneDrive, reformats, re-uploads.
"""
import os
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

ONEDRIVE_DIR = 'Documents/AI Tools/The Quarry'
TOKEN_URL    = 'https://login.microsoftonline.com/consumers/oauth2/v2.0/token'
GRAPH_URL    = 'https://graph.microsoft.com/v1.0'
COMPS_FILE   = Path('comps.xlsx')

FMT_DOLLARS = '$#,##0'
FMT_COMMA   = '#,##0'
FMT_PCT     = '0.0"%"'

# col index -> format for each tab (data rows only, skipping header row 1)
SALES_FORMATS  = {6: FMT_COMMA, 8: FMT_DOLLARS, 9: FMT_DOLLARS, 10: FMT_DOLLARS, 12: FMT_PCT}
LEASES_FORMATS = {6: FMT_COMMA}
LOANS_FORMATS  = {6: FMT_COMMA, 8: FMT_DOLLARS, 9: FMT_DOLLARS, 10: FMT_DOLLARS}


def _get_access_token():
    resp = requests.post(TOKEN_URL, data={
        'grant_type':    'refresh_token',
        'refresh_token': os.getenv('ONEDRIVE_REFRESH_TOKEN'),
        'client_id':     os.getenv('ONEDRIVE_CLIENT_ID'),
        'client_secret': os.getenv('ONEDRIVE_CLIENT_SECRET'),
        'scope':         'offline_access Files.ReadWrite',
    })
    resp.raise_for_status()
    return resp.json()['access_token']


def _download(token, filename):
    url  = f'{GRAPH_URL}/me/drive/root:/{ONEDRIVE_DIR}/{filename}:/content'
    resp = requests.get(url, headers={'Authorization': f'Bearer {token}'})
    resp.raise_for_status()
    path = Path(filename)
    path.write_bytes(resp.content)
    print(f'Downloaded {filename}')
    return path


def _upload(token, path):
    url  = f'{GRAPH_URL}/me/drive/root:/{ONEDRIVE_DIR}/{path.name}:/content'
    resp = requests.put(url, headers={
        'Authorization': f'Bearer {token}',
        'Content-Type':  'application/octet-stream',
    }, data=path.read_bytes())
    resp.raise_for_status()
    print(f'Uploaded {path.name}')


def _apply_formats(ws, fmt_map):
    count = 0
    for row in range(2, ws.max_row + 1):
        for col, fmt in fmt_map.items():
            ws.cell(row, col).number_format = fmt
        count += 1
    return count


def main():
    token = _get_access_token()

    _download(token, 'comps.xlsx')
    wb = load_workbook(COMPS_FILE)

    for sheet, fmt_map in [('Sales', SALES_FORMATS), ('Leases', LEASES_FORMATS), ('Loans', LOANS_FORMATS)]:
        if sheet in wb.sheetnames:
            n = _apply_formats(wb[sheet], fmt_map)
            print(f'  {sheet}: formatted {n} rows')

    wb.save(COMPS_FILE)
    _upload(token, COMPS_FILE)
    print('Done.')


if __name__ == '__main__':
    main()
