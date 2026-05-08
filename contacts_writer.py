import os
import anthropic
from openpyxl import Workbook

CONTACTS_COLS = ['Name', 'Title', 'Company', 'Role', 'City', 'State',
                 'Last Seen', 'First Seen', 'Appearances', 'Notes', 'Review Flag']

C_NAME        = 1
C_TITLE       = 2
C_COMPANY     = 3
C_ROLE        = 4
C_CITY        = 5
C_STATE       = 6
C_LAST_SEEN   = 7
C_FIRST_SEEN  = 8
C_APPEARANCES = 9
C_NOTES       = 10
C_REVIEW_FLAG = 11

_ai = None

def _client():
    global _ai
    if _ai is None:
        _ai = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
    return _ai


def _split_market(market):
    if not market:
        return None, None
    parts = market.rsplit(',', 1)
    city  = parts[0].strip() if parts else None
    state = parts[1].strip() if len(parts) > 1 else None
    return city, state


def _get_or_create_tab(wb, name, columns):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(columns)
    return ws


def _load_index(ws):
    index = {}
    for row_num in range(2, ws.max_row + 1):
        name    = str(ws.cell(row_num, C_NAME).value    or '').strip().lower()
        company = str(ws.cell(row_num, C_COMPANY).value or '').strip().lower()
        if name:
            index[(name, company)] = row_num
    return index


def _classify_people(people_data, narrative):
    if not people_data:
        return {}

    lines = '\n'.join(
        f'- Name: {name} | Raw title: {title or "unknown"} | Firm: {firm or "unknown"}'
        for name, title, firm in people_data
    )

    try:
        msg = _client().messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=1024,
            messages=[{
                'role': 'user',
                'content': (
                    f'Article context:\n{narrative}\n\n'
                    f'People mentioned:\n{lines}\n\n'
                    'For each person do three things:\n'
                    '1. Is this a CRE (commercial real estate) professional? Include brokers, '
                    'investors, developers, lenders, operators, asset managers, and executives '
                    'at real estate firms. Exclude politicians, government officials, residential '
                    'brokers/agents, project managers, engineers, architects, interior designers, '
                    'construction workers, and anyone outside CRE. '
                    'When in doubt, exclude.\n'
                    '2. Normalize their title to just the job function — strip any company/division '
                    'references (e.g. "CEO of JLL\'s Hotels and Hospitality, Americas" → "CEO").\n'
                    '3. Normalize their company — if the raw title contains a division or subsidiary, '
                    'use that as the company instead of the firm field '
                    '(e.g. "CEO of JLL\'s Hotels and Hospitality, Americas" → "JLL Hotels and Hospitality, Americas"). '
                    'Otherwise leave blank and the original firm will be used.\n\n'
                    'One line per person, pipe-delimited, exact format:\n'
                    'EXACT_NAME|YES_OR_NO|NORMALIZED_TITLE|NORMALIZED_COMPANY'
                ),
            }],
        )
        lookup = {n.lower(): (n, t, f) for n, t, f in people_data}
        result = {}
        for line in msg.content[0].text.strip().splitlines():
            parts = [p.strip() for p in line.split('|')]
            if len(parts) < 2:
                continue
            name_raw     = parts[0].lstrip('- ')
            is_cre       = parts[1].upper() == 'YES'
            norm_title   = parts[2] if len(parts) > 2 else ''
            norm_company = parts[3] if len(parts) > 3 else ''
            if name_raw.lower() in lookup:
                _, orig_title, orig_firm = lookup[name_raw.lower()]
                result[name_raw.lower()] = {
                    'cre':     is_cre,
                    'title':   norm_title or orig_title,
                    'company': norm_company or orig_firm,
                }
        return result
    except Exception:
        return {
            n.lower(): {'cre': True, 'title': t, 'company': f}
            for n, t, f in people_data
        }


def _note_entry(title):
    snippet = title[:80].rstrip()
    if len(title) > 80:
        snippet += '...'
    return snippet


def upsert_contacts(date_str: str, articles: list, wb: Workbook) -> int:
    ws = _get_or_create_tab(wb, 'Contacts', CONTACTS_COLS)
    index = _load_index(ws)
    new_count = 0

    ALLOWED_TX = {'sale', 'acquisition', 'lease', 'loan', 'refinance',
                  'development', 'construction', 'promotion'}
    SFR_TYPES  = {'single family', 'single-family', 'sfr', 'single family residential'}

    for article in articles:
        tx        = (article.get('transaction_type') or '').lower()
        cp        = article.get('companies_people') or []
        title     = article.get('title') or ''
        narrative = article.get('narrative') or ''
        city, state = _split_market(article.get('market'))

        if tx not in ALLOWED_TX:
            continue

        dp       = article.get('data_points') or {}
        raw_type = (dp.get('property_type') or '').lower()
        units    = dp.get('size_units')
        if raw_type in SFR_TYPES and (not units or units <= 1):
            continue

        all_people = [
            (person.get('name', '').strip(),
             person.get('title', '').strip(),
             entry.get('firm_name', '').strip())
            for entry in cp
            for person in (entry.get('people') or [])
            if person.get('name', '').strip()
        ]
        classifications = _classify_people(all_people, narrative)

        for entry in cp:
            role   = (entry.get('label') or '').upper()
            people = entry.get('people') or []

            for person in people:
                name = (person.get('name') or '').strip()
                if not name:
                    continue
                info = classifications.get(name.lower())
                if not info or not info['cre'] or not info['company']:
                    continue

                person_title = info['title']
                firm         = info['company']
                note         = _note_entry(title)
                key          = (name.lower(), firm.lower())

                if key in index:
                    row_num = index[key]
                    flagged = False

                    old_title = str(ws.cell(row_num, C_TITLE).value or '').strip()
                    old_role  = str(ws.cell(row_num, C_ROLE).value  or '').strip()
                    old_city  = str(ws.cell(row_num, C_CITY).value  or '').strip()
                    old_state = str(ws.cell(row_num, C_STATE).value or '').strip()

                    if person_title and person_title != old_title:
                        ws.cell(row_num, C_TITLE).value = person_title
                        flagged = True
                    if role and role != old_role:
                        ws.cell(row_num, C_ROLE).value = role
                        flagged = True
                    if city and city != old_city:
                        ws.cell(row_num, C_CITY).value = city
                        flagged = True
                    if state and state != old_state:
                        ws.cell(row_num, C_STATE).value = state
                        flagged = True

                    ws.cell(row_num, C_LAST_SEEN).value = date_str
                    appearances = int(ws.cell(row_num, C_APPEARANCES).value or 0) + 1
                    ws.cell(row_num, C_APPEARANCES).value = appearances
                    existing_notes = str(ws.cell(row_num, C_NOTES).value or '')
                    ws.cell(row_num, C_NOTES).value = (
                        existing_notes + ' | ' + note if existing_notes else note
                    )
                    if flagged:
                        ws.cell(row_num, C_REVIEW_FLAG).value = 'YES'

                else:
                    name_lower = name.lower()
                    same_name_exists = any(k[0] == name_lower for k in index)
                    review_flag = 'YES' if same_name_exists else ''

                    ws.append([
                        name,
                        person_title,
                        firm,
                        role,
                        city,
                        state,
                        date_str,
                        date_str,
                        1,
                        note,
                        review_flag,
                    ])
                    index[key] = ws.max_row
                    new_count += 1

    return new_count
