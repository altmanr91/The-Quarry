import os
import anthropic
from openpyxl import Workbook
from openpyxl.styles import PatternFill

YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

FMT_DOLLARS = '$#,##0'
FMT_COMMA   = '#,##0'
FMT_PCT     = '0.0"%"'

_ai = None

def _client():
    global _ai
    if _ai is None:
        _ai = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
    return _ai


def _fmt(ws, row, col_fmt_pairs):
    for col, fmt in col_fmt_pairs:
        ws.cell(row, col).number_format = fmt

SALES_COLS  = ['Date', 'Property Name', 'Address', 'City', 'State', 'Property Type',
               'Size (SF)', 'Units', 'Sale Price', '$/SF', '$/Unit',
               'Year Built', 'Occupancy %', 'Buyer', 'Seller', 'Broker', 'Lender',
               'Source', 'Link', 'Notes']

LEASES_COLS = ['Date', 'Property Name', 'Address', 'City', 'State', 'Property Type',
               'Size (SF)', 'Rent ($/SF/yr)', 'Tenants', 'Landlord', 'Tenant Rep', 'Landlord Broker',
               'Source', 'Link']

LOANS_COLS  = ['Date', 'Property Name', 'Address', 'City', 'State', 'Property Type',
               'Size (SF)', 'Units', 'Loan Amount', 'Loan/SF', 'Loan/Unit',
               'Borrower/Sponsor', 'Lender', 'Source', 'Link', 'Notes']

SFR_COLS    = ['Date', 'Address', 'City', 'State', 'Beds', 'Size (SF)',
               'Sale Price', '$/SF', 'Year Built',
               'Buyer', 'Seller', 'Broker', 'Lender',
               'Source', 'Link', 'Notes']

SALE_TYPES  = {'sale', 'acquisition'}
LEASE_TYPES = {'lease'}
LOAN_TYPES  = {'loan', 'refinance'}
SFR_TYPES   = {'single family', 'single-family', 'sfr', 'single family residential'}


def _get_or_create_tab(wb, name, columns):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(columns)
    return ws


def _firms_list(companies_people, *labels):
    labels_up = {l.upper() for l in labels}
    return [
        e['firm_name'] for e in companies_people
        if e.get('label', '').upper() in labels_up and e.get('firm_name')
    ]


def _filter_firms(firms, role, narrative):
    """Use Claude Haiku to filter to directly-involved firms when >4 are listed."""
    if len(firms) <= 4:
        return firms
    try:
        firm_list = '\n'.join(f'- {f}' for f in firms)
        msg = _client().messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=256,
            messages=[{
                'role': 'user',
                'content': (
                    f'Real estate transaction article:\n\n{narrative}\n\n'
                    f'Firms listed as {role}:\n{firm_list}\n\n'
                    f'Which of these firms were directly involved in this transaction as {role}? '
                    f'Return only their exact names, one per line, no other text.'
                ),
            }],
        )
        lines = [l.strip().lstrip('- ') for l in msg.content[0].text.strip().splitlines() if l.strip()]
        lookup = {f.lower(): f for f in firms}
        filtered = [lookup[l.lower()] for l in lines if l.lower() in lookup]
        return filtered if filtered else firms[:4]
    except Exception:
        return firms[:4]


def _firms(companies_people, *labels, narrative=''):
    lst = _firms_list(companies_people, *labels)
    role = labels[0].lower() if labels else 'participant'
    lst = _filter_firms(lst, role, narrative)
    return ', '.join(lst)


def _calc(numerator, denominator):
    if numerator and denominator and denominator > 0:
        return round(numerator / denominator, 2)
    return None


def _split_market(market):
    if not market:
        return None, None
    parts = market.rsplit(',', 1)
    city  = parts[0].strip() if parts else None
    state = parts[1].strip() if len(parts) > 1 else None
    return city, state


def _load_addresses(ws, addr_col_idx):
    result = {}
    for row_num in range(2, ws.max_row + 1):
        addr = str(ws.cell(row_num, addr_col_idx).value or '').strip().lower()
        if addr:
            result.setdefault(addr, []).append(row_num)
    return result


def _highlight_row(ws, row_num):
    for col in range(1, ws.max_column + 1):
        ws.cell(row_num, col).fill = YELLOW


def _check_duplicate(ws, addr_map, address, addr_col_idx):
    new_row    = ws.max_row
    addr_lower = (address or '').strip().lower()
    if not addr_lower:
        return
    if addr_lower in addr_map:
        _highlight_row(ws, new_row)
        for prev in addr_map[addr_lower]:
            _highlight_row(ws, prev)
        addr_map[addr_lower].append(new_row)
    else:
        addr_map[addr_lower] = [new_row]


def _purge_no_basis(wb: Workbook) -> int:
    removed = 0
    checks = [
        ('Sales',  9, 6),
        ('Leases', 7, None),
        ('Loans',  9, None),
        ('SFR',    7, None),
    ]
    for sheet_name, basis_col, type_col in checks:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(ws.max_row, 1, -1):
            no_basis = not ws.cell(row, basis_col).value
            no_type  = type_col and not ws.cell(row, type_col).value
            if no_basis or no_type:
                ws.delete_rows(row)
                removed += 1
    return removed


def append_articles(date_str: str, articles: list, wb: Workbook) -> dict:
    sales_ws  = _get_or_create_tab(wb, 'Sales',  SALES_COLS)
    leases_ws = _get_or_create_tab(wb, 'Leases', LEASES_COLS)
    loans_ws  = _get_or_create_tab(wb, 'Loans',  LOANS_COLS)
    sfr_ws    = _get_or_create_tab(wb, 'SFR',    SFR_COLS)

    sales_addrs  = _load_addresses(sales_ws,  3)
    leases_addrs = _load_addresses(leases_ws, 3)
    loans_addrs  = _load_addresses(loans_ws,  3)
    sfr_addrs    = _load_addresses(sfr_ws,    2)

    counts = {'sales': 0, 'leases': 0, 'loans': 0, 'sfr': 0}

    for article in articles:
        tx = (article.get('transaction_type') or '').lower()
        if tx not in SALE_TYPES | LEASE_TYPES | LOAN_TYPES:
            continue

        dp        = article.get('data_points') or {}
        cp        = article.get('companies_people') or []
        narrative = article.get('narrative') or ''
        addr      = dp.get('address') or ''
        name      = dp.get('property_name') or (addr.split(',')[0].strip() if addr else None)
        ptype     = dp.get('property_type', '').title() if dp.get('property_type') else None
        city, state = _split_market(article.get('market'))
        raw_ptype   = (dp.get('property_type') or '').lower()
        units       = dp.get('size_units')
        is_sfr      = raw_ptype in SFR_TYPES and (not units or units <= 1)

        if tx in SALE_TYPES and is_sfr:
            if not dp.get('sale_price'):
                continue
            sfr_ws.append([
                date_str,
                addr,
                city,
                state,
                dp.get('size_beds'),
                dp.get('size_sf'),
                dp.get('sale_price'),
                _calc(dp.get('sale_price'), dp.get('size_sf')),
                dp.get('year_built'),
                _firms(cp, 'BUYER',                        narrative=narrative),
                _firms(cp, 'SELLER',                       narrative=narrative),
                _firms(cp, 'SELLER BROKER', 'BUYER BROKER', narrative=narrative),
                _firms(cp, 'LENDER',                       narrative=narrative),
                article.get('source'),
                article.get('link'),
                article.get('financing'),
            ])
            _check_duplicate(sfr_ws, sfr_addrs, addr, 2)
            _fmt(sfr_ws, sfr_ws.max_row, [
                (6, FMT_COMMA),
                (7, FMT_DOLLARS),
                (8, FMT_DOLLARS),
            ])
            counts['sfr'] += 1

        elif tx in SALE_TYPES:
            if not dp.get('sale_price') or not dp.get('property_type'):
                continue
            sales_ws.append([
                date_str,
                name,
                addr,
                city,
                state,
                ptype,
                dp.get('size_sf'),
                dp.get('size_units'),
                dp.get('sale_price'),
                _calc(dp.get('sale_price'), dp.get('size_sf')),
                _calc(dp.get('sale_price'), dp.get('size_units')),
                dp.get('year_built'),
                dp.get('occupancy'),
                _firms(cp, 'BUYER',                        narrative=narrative),
                _firms(cp, 'SELLER',                       narrative=narrative),
                _firms(cp, 'SELLER BROKER', 'BUYER BROKER', narrative=narrative),
                _firms(cp, 'LENDER',                       narrative=narrative),
                article.get('source'),
                article.get('link'),
                article.get('financing'),
            ])
            _check_duplicate(sales_ws, sales_addrs, addr, 3)
            _fmt(sales_ws, sales_ws.max_row, [
                (7, FMT_COMMA),
                (9, FMT_DOLLARS),
                (10, FMT_DOLLARS),
                (11, FMT_DOLLARS),
                (13, FMT_PCT),
            ])
            counts['sales'] += 1

        elif tx in LEASE_TYPES:
            if not dp.get('size_sf'):
                continue
            tenants = ', '.join(article.get('tenants') or [])
            leases_ws.append([
                date_str,
                name,
                addr,
                city,
                state,
                ptype,
                dp.get('size_sf'),
                dp.get('rental_rate'),
                tenants,
                _firms(cp, 'LANDLORD', 'OWNER', narrative=narrative),
                _firms(cp, 'TENANT REP',         narrative=narrative),
                _firms(cp, 'LEASING AGENT',       narrative=narrative),
                article.get('source'),
                article.get('link'),
            ])
            _check_duplicate(leases_ws, leases_addrs, addr, 3)
            _fmt(leases_ws, leases_ws.max_row, [
                (7, FMT_COMMA),
                (8, FMT_DOLLARS),
            ])
            counts['leases'] += 1

        elif tx in LOAN_TYPES:
            if not dp.get('loan_amount'):
                continue
            loans_ws.append([
                date_str,
                name,
                addr,
                city,
                state,
                ptype,
                dp.get('size_sf'),
                dp.get('size_units'),
                dp.get('loan_amount'),
                _calc(dp.get('loan_amount'), dp.get('size_sf')),
                _calc(dp.get('loan_amount'), dp.get('size_units')),
                _firms(cp, 'SPONSOR', 'DEVELOPER/SPONSOR', narrative=narrative),
                _firms(cp, 'LENDER',                       narrative=narrative),
                article.get('source'),
                article.get('link'),
                article.get('financing'),
            ])
            _check_duplicate(loans_ws, loans_addrs, addr, 3)
            _fmt(loans_ws, loans_ws.max_row, [
                (7, FMT_COMMA),
                (9, FMT_DOLLARS),
                (10, FMT_DOLLARS),
                (11, FMT_DOLLARS),
            ])
            counts['loans'] += 1

    return counts
